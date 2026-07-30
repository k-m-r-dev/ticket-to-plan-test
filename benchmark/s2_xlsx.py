#!/usr/bin/env python3
"""Excel front-end for ABP S2 human scoring (+ agent draft merge).

Commands:
  emit          — rebuild S2_SCORECARD.xlsx + opaque plan copies from mapping_key
  merge-drafts  — fold human_sheets/drafts/*.json into the scorecard (review queue)
  apply         — write scorecard answers into human_sheets/*.json (official path)
  status        — draft / review / complete counts

Official S2 = what you leave in the scorecard after reviewing needs_review rows,
then apply + human_sheet.py ingest.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BENCH = Path(__file__).resolve().parent
ROOT = BENCH.parent

FONT = "Arial"


def paths(protocol: str) -> dict[str, Path]:
    hs = BENCH / "runs" / protocol / "human_sheets"
    return {
        "hs": hs,
        "xlsx": hs / "S2_SCORECARD.xlsx",
        "plans": hs / "plans",
        "drafts": hs / "drafts",
        "key": hs / "mapping_key.json",
    }


def _as_bool(val) -> bool | None:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if s in {"TRUE", "YES", "Y", "1"}:
        return True
    if s in {"FALSE", "NO", "N", "0"}:
        return False
    raise ValueError(f"Expected TRUE/FALSE, got {val!r}")


def _tf(val: bool | None) -> str:
    if val is None:
        return ""
    return "TRUE" if val else "FALSE"


def load_key(protocol: str) -> dict:
    return json.loads(paths(protocol)["key"].read_text(encoding="utf-8"))


def ensure_opaque_plans(protocol: str) -> int:
    p = paths(protocol)
    key = load_key(protocol)
    if p["plans"].exists():
        shutil.rmtree(p["plans"])
    p["plans"].mkdir(parents=True)
    for blind_id, info in key.items():
        dest = p["plans"] / blind_id
        src = ROOT / info["run_dir"] / "artifacts"
        dest.mkdir(parents=True)
        if src.is_dir():
            for f in src.rglob("*"):
                if f.is_file():
                    rel = f.relative_to(src)
                    out = dest / rel
                    out.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(f, out)
        else:
            (dest / "MISSING_ARTIFACTS.txt").write_text(f"No artifacts at {src}\n")
    return len(key)


def _styles():
    return {
        "title": Font(name=FONT, size=16, bold=True, color="1A1F2E"),
        "header": Font(name=FONT, size=11, bold=True, color="FFFFFF"),
        "label": Font(name=FONT, size=11, bold=True, color="1A1F2E"),
        "body": Font(name=FONT, size=11, color="1A1F2E"),
        "input": Font(name=FONT, size=11, color="0000FF"),
        "muted": Font(name=FONT, size=10, color="5C6578"),
        "header_fill": PatternFill("solid", fgColor="0F766E"),
        "yellow": PatternFill("solid", fgColor="FFFF99"),
        "flag": PatternFill("solid", fgColor="FDE68A"),
        "soft": PatternFill("solid", fgColor="F4F6F8"),
        "thin": Border(
            left=Side(style="thin", color="D0D5DD"),
            right=Side(style="thin", color="D0D5DD"),
            top=Side(style="thin", color="D0D5DD"),
            bottom=Side(style="thin", color="D0D5DD"),
        ),
        "wrap": Alignment(wrap_text=True, vertical="center"),
    }


def _build_workbook(protocol: str, row_data: dict) -> Path:
    """row_data[blind_id] = {depth_ok, guardrails_ok, notes, depth_needs_review, ... items: {id: {...}}}"""
    st = _styles()
    p = paths(protocol)
    key = load_key(protocol)
    ordered = sorted(key.items(), key=lambda kv: (kv[1]["fixture"], kv[0]))

    wb = Workbook()
    g = wb.active
    g.title = "Guide"
    g.sheet_view.showGridLines = False
    g["A1"] = "S2 scoring — draft + human review"
    g["A1"].font = st["title"]
    g.merge_cells("A1:B1")

    steps = [
        ("A. Optional agent draft", "From your terminal: python3 benchmark/run_s2_draft.py --keep-going"),
        ("B. Merge drafts", "python3 benchmark/s2_xlsx.py merge-drafts"),
        ("C. Review queue", "On Items/Runs, filter needs_review=TRUE. Fill blank Present / flags. Clear needs_review when done."),
        ("D. Spot-check", "Sample some needs_review=FALSE rows — you remain the authority."),
        ("E. Save workbook", "Keep filename S2_SCORECARD.xlsx"),
        ("F. Apply + ingest", "python3 benchmark/s2_xlsx.py apply && python3 benchmark/human_sheet.py ingest --protocol abp-v1"),
    ]
    g["A3"] = "Flow"
    g["A3"].font = st["label"]
    for i, (t, d) in enumerate(steps, start=4):
        g[f"A{i}"] = t
        g[f"A{i}"].font = st["label"]
        g[f"B{i}"] = d
        g[f"B{i}"].font = st["body"]
        g[f"B{i}"].alignment = st["wrap"]
        g.row_dimensions[i].height = 34

    g["A11"] = "Rules"
    g["A11"].font = st["label"]
    for i, rule in enumerate(
        [
            "Yellow/blue cells = your inputs (official after apply).",
            "Amber needs_review = agent was unsure OR you still need to look.",
            "Agent draft is NOT official S2 until you apply after review.",
            "Do NOT open mapping_key.json or ANSWER_KEY.md while scoring.",
            "Plan folders under human_sheets/plans/<Blind ID>/ are opaque copies (no arm name).",
        ],
        start=12,
    ):
        g[f"A{i}"] = "•"
        g[f"B{i}"] = rule
        g[f"B{i}"].font = st["body"]
        g[f"B{i}"].alignment = st["wrap"]
        g.row_dimensions[i].height = 28

    g.column_dimensions["A"].width = 28
    g.column_dimensions["B"].width = 92

    # Runs
    r = wb.create_sheet("Runs")
    run_headers = [
        "Blind ID",
        "Fixture",
        "Plan folder (open this)",
        "depth_ok",
        "guardrails_ok",
        "depth_needs_review",
        "guardrails_needs_review",
        "agent_note",
        "Notes (human)",
        "Items answered",
        "Items total",
        "Review queue",
        "Progress %",
    ]
    for col, h in enumerate(run_headers, 1):
        cell = r.cell(1, col, h)
        cell.font = st["header"]
        cell.fill = st["header_fill"]
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    r.add_data_validation(dv_bool)

    for i, (blind_id, info) in enumerate(ordered, start=2):
        sheet = json.loads((p["hs"] / f"{blind_id}.json").read_text(encoding="utf-8"))
        data = row_data.get(blind_id, {})
        n_items = len(sheet["items"])
        r.cell(i, 1, blind_id).font = st["body"]
        r.cell(i, 2, info["fixture"]).font = st["body"]
        r.cell(i, 3, f"benchmark/runs/{protocol}/human_sheets/plans/{blind_id}").font = st["muted"]

        depth = data.get("depth_ok", sheet.get("depth_ok"))
        guard = data.get("guardrails_ok", sheet.get("guardrails_ok"))
        dnr = data.get("depth_needs_review", False)
        gnr = data.get("guardrails_needs_review", False)

        pairs = [(4, depth, False), (5, guard, False), (6, dnr, True), (7, gnr, True)]
        for col, val, is_flag in pairs:
            if is_flag:
                c = r.cell(i, col, "TRUE" if val else "FALSE")
                c.fill = st["flag"] if val else st["yellow"]
            else:
                c = r.cell(i, col, _tf(val))
                c.fill = st["yellow"]
            c.font = st["input"]
            c.border = st["thin"]
            c.alignment = Alignment(horizontal="center", vertical="center")

        an = r.cell(i, 8, data.get("agent_note") or data.get("summary_note") or "")
        an.font = st["muted"]
        an.fill = st["flag"] if an.value else st["soft"]
        an.border = st["thin"]
        an.alignment = st["wrap"]

        notes = r.cell(i, 9, data.get("notes", sheet.get("notes") or ""))
        notes.font = st["input"]
        notes.fill = st["yellow"]
        notes.border = st["thin"]

        r.cell(
            i,
            10,
            f'=COUNTIFS(Items!A:A,A{i},Items!E:E,"TRUE")+COUNTIFS(Items!A:A,A{i},Items!E:E,"FALSE")',
        ).font = st["body"]
        r.cell(i, 11, n_items).font = st["body"]
        r.cell(
            i,
            12,
            f'=COUNTIFS(Items!A:A,A{i},Items!F:F,"TRUE")+IF(OR(F{i}="TRUE",G{i}="TRUE"),1,0)',
        ).font = st["body"]
        prog = r.cell(i, 13, f'=IF(K{i}=0,"",J{i}/K{i})')
        prog.font = st["body"]
        prog.number_format = "0%"

        for col in range(1, 14):
            r.cell(i, col).border = st["thin"]
            if col not in (4, 5, 6, 7, 8, 9) and i % 2 == 0:
                r.cell(i, col).fill = st["soft"]

    dv_bool.add(f"D2:G{len(ordered) + 1}")
    for i, w in enumerate([12, 10, 48, 11, 14, 16, 18, 36, 24, 12, 10, 12, 11], 1):
        r.column_dimensions[get_column_letter(i)].width = w
    r.row_dimensions[1].height = 34
    r.freeze_panes = "A2"
    r.auto_filter.ref = f"A1:M{len(ordered) + 1}"

    # Items
    it = wb.create_sheet("Items")
    item_headers = [
        "Blind ID",
        "Fixture",
        "Item ID",
        "What to look for in the plan",
        "Present",
        "needs_review",
        "agent_note",
        "evidence",
    ]
    for col, h in enumerate(item_headers, 1):
        cell = it.cell(1, col, h)
        cell.font = st["header"]
        cell.fill = st["header_fill"]
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    dv_present = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    it.add_data_validation(dv_present)

    row = 2
    for blind_id, info in ordered:
        sheet = json.loads((p["hs"] / f"{blind_id}.json").read_text(encoding="utf-8"))
        draft_items = row_data.get(blind_id, {}).get("items", {})
        for item in sheet["items"]:
            ditem = draft_items.get(item["id"], {})
            present = ditem.get("present", item.get("present"))
            needs = bool(ditem.get("needs_review", False))
            # if draft left null with needs_review, present stays blank
            if ditem.get("needs_review") and ditem.get("present") is None:
                present = None

            it.cell(row, 1, blind_id).font = st["body"]
            it.cell(row, 2, sheet["fixture"]).font = st["body"]
            it.cell(row, 3, item["id"]).font = st["muted"]
            it.cell(row, 4, item.get("text") or item["id"]).font = st["body"]
            it.cell(row, 4).alignment = st["wrap"]

            c = it.cell(row, 5, _tf(present))
            c.font = st["input"]
            c.fill = st["yellow"]
            c.alignment = Alignment(horizontal="center", vertical="center")

            nr = it.cell(row, 6, "TRUE" if needs else "FALSE")
            nr.font = st["input"]
            nr.fill = st["flag"] if needs else st["yellow"]
            nr.alignment = Alignment(horizontal="center", vertical="center")

            note = it.cell(row, 7, ditem.get("note") or "")
            note.font = st["muted"]
            note.fill = st["flag"] if needs else st["soft"]
            note.alignment = st["wrap"]

            ev = it.cell(row, 8, ditem.get("evidence") or "")
            ev.font = st["muted"]
            ev.alignment = st["wrap"]

            for col in range(1, 9):
                it.cell(row, col).border = st["thin"]
                if col not in (5, 6, 7) and row % 2 == 0:
                    it.cell(row, col).fill = st["soft"]
            row += 1

    last = row - 1
    dv_present.add(f"E2:F{last}")
    for letter, w in zip("ABCDEFGH", [12, 10, 16, 48, 10, 14, 36, 40]):
        it.column_dimensions[letter].width = w
    it.row_dimensions[1].height = 34
    it.freeze_panes = "A2"
    it.auto_filter.ref = f"A1:H{last}"

    leg = wb.create_sheet("Legend")
    leg["A1"] = "Legend"
    leg["A1"].font = st["title"]
    leg["A3"] = "Yellow + blue"
    leg["A3"].font = st["input"]
    leg["A3"].fill = st["yellow"]
    leg["B3"] = "Your official inputs"
    leg["A4"] = "Amber needs_review / agent_note"
    leg["A4"].fill = st["flag"]
    leg["B4"] = "Review queue — fill Present, then set needs_review=FALSE"
    leg["A6"] = "Filter tip"
    leg["A6"].font = st["label"]
    leg["B6"] = "On Items: filter needs_review column = TRUE to see only the queue."
    leg.column_dimensions["A"].width = 28
    leg.column_dimensions["B"].width = 70

    wb.save(p["xlsx"])
    return p["xlsx"]


def _empty_row_data(protocol: str) -> dict:
    """Seed from existing JSON sheets (no drafts)."""
    p = paths(protocol)
    key = load_key(protocol)
    out = {}
    for blind_id in key:
        sheet = json.loads((p["hs"] / f"{blind_id}.json").read_text(encoding="utf-8"))
        out[blind_id] = {
            "depth_ok": sheet.get("depth_ok"),
            "guardrails_ok": sheet.get("guardrails_ok"),
            "depth_needs_review": False,
            "guardrails_needs_review": False,
            "agent_note": "",
            "notes": sheet.get("notes") or "",
            "items": {
                it["id"]: {
                    "present": it.get("present"),
                    "needs_review": False,
                    "note": "",
                    "evidence": "",
                }
                for it in sheet["items"]
            },
        }
    return out


def _merge_draft_into(row_data: dict, draft: dict) -> None:
    blind_id = draft["blind_id"]
    slot = row_data.setdefault(
        blind_id,
        {
            "depth_ok": None,
            "guardrails_ok": None,
            "depth_needs_review": False,
            "guardrails_needs_review": False,
            "agent_note": "",
            "notes": "",
            "items": {},
        },
    )
    # depth / guardrails
    if draft.get("depth_needs_review"):
        slot["depth_ok"] = None
        slot["depth_needs_review"] = True
    else:
        slot["depth_ok"] = draft.get("depth_ok")
        slot["depth_needs_review"] = False
    if draft.get("guardrails_needs_review"):
        slot["guardrails_ok"] = None
        slot["guardrails_needs_review"] = True
    else:
        slot["guardrails_ok"] = draft.get("guardrails_ok")
        slot["guardrails_needs_review"] = False

    notes = []
    if draft.get("depth_note"):
        notes.append(f"depth: {draft['depth_note']}")
    if draft.get("guardrails_note"):
        notes.append(f"guardrails: {draft['guardrails_note']}")
    if draft.get("summary_note"):
        notes.append(draft["summary_note"])
    slot["agent_note"] = " | ".join(notes)

    for it in draft.get("items", []):
        iid = it["id"]
        needs = bool(it.get("needs_review"))
        present = it.get("present")
        if needs:
            present = None
        slot["items"][iid] = {
            "present": present,
            "needs_review": needs,
            "note": it.get("note") or "",
            "evidence": it.get("evidence") or "",
        }


def emit(protocol: str = "abp-v1") -> None:
    n = ensure_opaque_plans(protocol)
    xlsx = _build_workbook(protocol, _empty_row_data(protocol))
    print(f"emitted {xlsx} (opaque plans={n})")


def merge_drafts(protocol: str = "abp-v1") -> None:
    p = paths(protocol)
    if not p["drafts"].is_dir():
        raise SystemExit(f"No drafts dir at {p['drafts']}. Run run_s2_draft.py first.")
    row_data = _empty_row_data(protocol)
    merged = 0
    for path in sorted(p["drafts"].glob("*.json")):
        draft = json.loads(path.read_text(encoding="utf-8"))
        if draft.get("blind_id") not in row_data:
            print(f"warning: skip unknown draft {path.name}", file=sys.stderr)
            continue
        _merge_draft_into(row_data, draft)
        merged += 1
    ensure_opaque_plans(protocol)
    xlsx = _build_workbook(protocol, row_data)
    # also write draft fields into JSON sheets as draft-only sidecar? keep official JSON clean until apply
    review_items = sum(
        1
        for b in row_data.values()
        for it in b["items"].values()
        if it.get("needs_review")
    )
    review_runs = sum(
        1
        for b in row_data.values()
        if b.get("depth_needs_review") or b.get("guardrails_needs_review")
    )
    print(f"merged {merged} drafts -> {xlsx}")
    print(f"review queue: {review_items} checklist rows, {review_runs} runs with flag review")


def apply(protocol: str = "abp-v1") -> None:
    p = paths(protocol)
    if not p["xlsx"].is_file():
        raise SystemExit(f"Missing {p['xlsx']}. Run emit or merge-drafts first.")

    wb = load_workbook(p["xlsx"], data_only=False)
    key = load_key(protocol)

    runs = {}
    for row in wb["Runs"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        blind_id = str(row[0]).strip()
        if blind_id not in key:
            continue
        runs[blind_id] = {
            "depth_ok": _as_bool(row[3]),
            "guardrails_ok": _as_bool(row[4]),
            # cols: 5 depth_needs_review, 6 guardrails_needs_review, 7 agent_note, 8 human notes
            "notes": (row[8] or "") if len(row) > 8 and row[8] is not None else "",
        }

    presents: dict[str, dict[str, bool | None]] = {b: {} for b in key}
    still_flagged = 0
    for row in wb["Items"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        blind_id = str(row[0]).strip()
        item_id = str(row[2]).strip()
        if blind_id not in key:
            continue
        needs = _as_bool(row[5]) if len(row) > 5 else False
        if needs:
            still_flagged += 1
            # leave present as-is in JSON only if human filled it; still allow apply of filled Present
        presents[blind_id][item_id] = _as_bool(row[4])

    if still_flagged:
        print(
            f"warning: {still_flagged} Items rows still have needs_review=TRUE "
            "(apply will still write any Present values you set)",
            file=sys.stderr,
        )

    updated = 0
    complete = 0
    for blind_id in key:
        path = p["hs"] / f"{blind_id}.json"
        sheet = json.loads(path.read_text(encoding="utf-8"))
        run = runs.get(blind_id, {})
        sheet["depth_ok"] = run.get("depth_ok")
        sheet["guardrails_ok"] = run.get("guardrails_ok")
        sheet["notes"] = run.get("notes", sheet.get("notes") or "")
        for item in sheet["items"]:
            if item["id"] in presents.get(blind_id, {}):
                item["present"] = presents[blind_id][item["id"]]
        path.write_text(json.dumps(sheet, indent=2) + "\n", encoding="utf-8")
        updated += 1
        answered = sum(1 for i in sheet["items"] if i.get("present") is not None)
        if (
            answered == len(sheet["items"])
            and sheet.get("depth_ok") is not None
            and sheet.get("guardrails_ok") is not None
        ):
            complete += 1

    print(f"applied {updated} sheets ({complete} complete) from {p['xlsx']}")
    print("next: python3 benchmark/human_sheet.py ingest --protocol", protocol)


def status(protocol: str = "abp-v1") -> None:
    p = paths(protocol)
    key = load_key(protocol)
    drafts = 0
    if p["drafts"].is_dir():
        drafts = sum(1 for _ in p["drafts"].glob("*.json"))
    complete = partial = empty = 0
    for blind_id in key:
        sheet = json.loads((p["hs"] / f"{blind_id}.json").read_text(encoding="utf-8"))
        answered = sum(1 for i in sheet["items"] if i.get("present") is not None)
        flags = sheet.get("depth_ok") is not None and sheet.get("guardrails_ok") is not None
        if answered == len(sheet["items"]) and flags:
            complete += 1
        elif answered or flags:
            partial += 1
        else:
            empty += 1
    print(
        f"S2 status: drafts={drafts}/{len(key)} | "
        f"json complete={complete} partial={partial} empty={empty}"
    )


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "command",
        choices=["emit", "merge-drafts", "apply", "status"],
    )
    ap.add_argument("--protocol", default="abp-v1")
    args = ap.parse_args()
    if args.command == "emit":
        emit(args.protocol)
    elif args.command == "merge-drafts":
        merge_drafts(args.protocol)
    elif args.command == "apply":
        apply(args.protocol)
    else:
        status(args.protocol)


if __name__ == "__main__":
    main()
